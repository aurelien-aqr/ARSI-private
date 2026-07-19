"""Report + spreadsheet exports for a finished job (docs/SPEC.md)."""
import html
import json
from datetime import datetime
from io import BytesIO


def report_md(data: dict) -> str:
    cfg, s = data["config"], data["summary"]
    L = [f"# ARSI Studio report — {data['job_id']}", "",
         f"**Status:** {data['status']}  ",
         f"**Pipeline:** `{cfg['script']}` · **Model:** `{cfg['model']}` · "
         f"**Prompt:** {cfg['prompt_name']}  ",
         f"**Mask:** {cfg.get('mask') or 'none'} · **Reference:** {cfg.get('reference') or 'n/a'}  ",
         f"**Started:** {data['started']} · **Finished:** {data['finished']} · "
         f"**Wall:** {s['wall_seconds']} s", "",
         "## Summary", "",
         f"- Frames: **{s['n_frames']}** ({s['n_ok']} ok, {s['n_failed']} failed)",
         f"- Anomalous frames: **{s['n_anomalous']}**", "",
         "## Per-frame results", "",
         "| frame | status | anomaly | detections | attempts | s |",
         "|---|---|---|---|---|---|"]
    for f in data["frames"]:
        dets = "; ".join(f"{d['label']} [{d['type']}]" for d in f["detections"]) or "—"
        anom = {True: "**YES**", False: "no"}.get(f["anomaly"], "—")
        L.append(f"| {f['frame_id']} | {f['status']} | {anom} | {dets} | "
                 f"{f['attempts']} | {f['seconds']} |")
    if cfg.get("prompt"):
        L += ["", "## Prompt", "", "```", cfg["prompt"], "```"]
    return "\n".join(L) + "\n"


def report_html(data: dict) -> str:
    body = []
    cfg, s = data["config"], data["summary"]
    body.append(f"<h1>ARSI Studio report — {html.escape(data['job_id'])}</h1>")
    body.append(f"<p><b>Status:</b> {data['status']} · <b>Pipeline:</b> "
                f"{html.escape(cfg['script'])} · <b>Model:</b> "
                f"{html.escape(str(cfg['model']))} · <b>Prompt:</b> "
                f"{html.escape(cfg['prompt_name'])} · <b>Mask:</b> "
                f"{html.escape(str(cfg.get('mask') or 'none'))}</p>")
    body.append("<div class='cards'>"
                f"<div class='card'><b>{s['n_frames']}</b><span>frames</span></div>"
                f"<div class='card red'><b>{s['n_anomalous']}</b><span>anomalous</span></div>"
                f"<div class='card'><b>{s['n_failed']}</b><span>failed</span></div>"
                f"<div class='card'><b>{s['wall_seconds']}s</b><span>wall clock</span></div>"
                "</div>")
    body.append("<table><tr><th>frame</th><th>status</th><th>anomaly</th>"
                "<th>detections</th><th>attempts</th><th>s</th></tr>")
    for f in data["frames"]:
        dets = "<br>".join(f"{html.escape(d['label'])} <i>[{d['type']}]</i>"
                           for d in f["detections"]) or "—"
        anom = {True: "<b class='yes'>YES</b>", False: "no"}.get(f["anomaly"], "—")
        cls = " class='failed'" if f["status"] == "failed" else ""
        body.append(f"<tr{cls}><td>{html.escape(f['frame_id'])}</td>"
                    f"<td>{f['status']}</td><td>{anom}</td><td>{dets}</td>"
                    f"<td>{f['attempts']}</td><td>{f['seconds']}</td></tr>")
    body.append("</table>")
    style = ("body{font-family:system-ui;margin:32px auto;max-width:960px;color:#1c2128}"
             "table{border-collapse:collapse;width:100%;font-size:14px}"
             "td,th{border:1px solid #d5dae1;padding:6px 10px;text-align:left;"
             "vertical-align:top}th{background:#f0f2f5}"
             ".yes{color:#c0392b}.failed{background:#fdf3f2}"
             ".cards{display:flex;gap:12px;margin:18px 0}"
             ".card{border:1px solid #d5dae1;border-radius:10px;padding:12px 18px;"
             "display:flex;flex-direction:column}.card b{font-size:22px}"
             ".card span{font-size:12px;color:#68707c}.card.red b{color:#c0392b}")
    return (f"<!doctype html><meta charset='utf-8'><title>{html.escape(data['job_id'])}"
            f"</title><style>{style}</style>" + "".join(body))


def results_xlsx(data: dict, review: dict = None, metrics: dict = None) -> bytes:
    """Rows in the spirit of ARSI_results_EN.xlsx: one row per frame. When a
    review exists, each row carries the human verdict (Correctness TP/FP/TN/FN
    per the supervisor's any-miss-is-FN rule) and a Review sheet holds the
    aggregated metrics — the by-hand grid, generated."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Frames"
    cfg = data["config"]
    correctness = (metrics or {}).get("correctness", {})
    rframes = (review or {}).get("frames", {})
    header = ["ID", "Date", "Task", "Model", "Test image", "Reference image",
              "Prompt", "Model response", "Anomaly (YES/NO)", "Anomaly type",
              "Detections", "Inference time [s]", "Status", "Note"]
    if review is not None:
        header += ["Correctness", "TP boxes", "FP boxes", "Missed (FN)", "Reviewed"]
    ws.append(header)
    date = (data.get("started") or "")[:10] or datetime.now().strftime("%Y-%m-%d")
    for i, f in enumerate(data["frames"]):
        types = ",".join(sorted({d["type"] for d in f["detections"]})) or ""
        dets = "; ".join(f"{d['label']} {d['bbox'] or ''}".strip()
                         for d in f["detections"])
        row = [i, date, cfg["script"], cfg["model"], f["image"],
               cfg.get("reference") or "", cfg["prompt_name"],
               (f["raw_response"] or "")[:900],
               {True: "YES", False: "NO"}.get(f["anomaly"], ""),
               types, dets, f["seconds"], f["status"], f.get("error") or ""]
        if review is not None:
            e = rframes.get(f["frame_id"], {})
            verdicts = e.get("verdicts", {})
            missed = "; ".join(f"{m['label']} {m['bbox']}" for m in e.get("missed", []))
            row += [correctness.get(f["frame_id"], ""),
                    sum(1 for v in verdicts.values() if v == "tp"),
                    sum(1 for v in verdicts.values() if v == "fp"),
                    missed, "yes" if e.get("done") else ""]
        ws.append(row)
    ws2 = wb.create_sheet("Summary")
    for k, v in {**data["summary"], "job_id": data["job_id"],
                 "status": data["status"], "mask": cfg.get("mask") or "none"}.items():
        ws2.append([k, json.dumps(v) if isinstance(v, (dict, list)) else v])
    if metrics:
        ws3 = wb.create_sheet("Review")
        ws3.append(["reviewed frames", metrics["progress"]["n_done"],
                    "of", metrics["progress"]["n_frames"]])
        ws3.append([])
        for k in ("tp", "fp", "fn", "precision", "recall", "f1"):
            ws3.append(["objects." + k, metrics["objects"][k]])
        for k in ("TP", "FP", "TN", "FN", "accuracy", "precision", "recall",
                  "specificity", "f1"):
            ws3.append(["frames." + k, metrics["frames"][k]])
        ws3.append([])
        ws3.append(["type", "tp", "fp", "fn", "recall"])
        for t, d in sorted(metrics["per_type"].items()):
            ws3.append([t, d["tp"], d["fp"], d["fn"], d["recall"]])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

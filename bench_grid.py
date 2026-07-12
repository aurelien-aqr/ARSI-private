#!/usr/bin/env python3
# =============================================================================
#  ARSI-VLM - bench_grid.py
#  Model x task x image benchmark grid for the ARSI results spreadsheet.
#
#  Runs the SAME prompts as vlm_01 / vlm_02 / vlm_03 (imported from those
#  modules, so there is a single source of truth) plus the ORIGINAL study's
#  Task-3 variant (zones of interest given TO the model, template from the
#  'Example Prompts' sheet, zones from benchmark/zones_tram_1762.json), against
#  every requested Ollama model, and records one spreadsheet row per call:
#  timing, full response, and a heuristic Anomaly YES/NO. Correctness / Rating
#  stay empty on purpose - judging against ground truth is the human's call.
#
#  Output: results/grid_results.csv (appended after EVERY call, so the grid is
#  crash-safe and RESUMABLE: existing (model, task, image, prompt) combinations
#  are skipped on re-run). Use --to-xlsx to also append the rows into a copy of
#  ARSI_results_EN.xlsx (a timestamped .bak of the workbook is written first).
#
#  Examples (from the repository root):
#    python bench_grid.py --dry-run
#    python bench_grid.py --models openbmb/minicpm-v4.6,llama3.2-vision:11b
#    python bench_grid.py --tasks 1,2 --models qwen3.5:9b
#    python bench_grid.py --to-xlsx ../ARSI_results_EN.xlsx
# =============================================================================

import sys, csv, json, time, re, shutil, argparse, hashlib
from datetime import date
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(REPO_ROOT))
import ollama
import vlm_01_single_image as v1
import vlm_02_reference_compare as v2
import vlm_03_bounding_box as v3

# =============================================================================
#  USER CONFIG
# =============================================================================
STUDENT = "Alquier A"

# Models to sweep (Ollama names). The first four are the still-untested pulls;
# the last four already have manual rows in the spreadsheet - the grid skips
# nothing by itself, so trim this list (or use --models) to control cost.
MODELS = [
    "openbmb/minicpm-v4.6",
    "haervwe/GLM-4.6V-Flash-9B",
    "hf.co/mradermacher/Cosmos-Reason2-8B-GGUF:Q4_K_M",
    "llama3.2-vision:11b",
    "qwen3-vl:8b-instruct",
    "qwen2.5vl:7b",
    "qwen3.5:9b",
    "blaifa/InternVL3_5:8b",
]

# Fixed image sets so every model answers the SAME questions.
TASK1_IMAGES = [                                   # single raw frame
    "data/raw/tram_1762_v1_f0001.jpg",             # clean (expected NO)
    "data/raw/tram_1762_v2_f0037.jpg",             # 4 forgotten objects
    "data/anomalies/tram_anomalies_4.png",         # big graffiti wall
    "data/anomalies/tram_anomalies_11.png",        # anomalies + 3 people
]
TASK2_PAIRS = [                                    # (reference, inspection)
    ("data/reference/tram_1762_v1_f0227_masked_reference.jpg",
     "data/masked/tram_1762_v2_f0032_masked.jpg"),
    ("data/reference/tram_1762_v1_f0227_masked_reference.jpg",
     "data/anomalies/tram_1762_v2_f0037_masked.jpg"),
    ("data/reference/tram_1762_v1_f0227_masked_reference.jpg",
     "data/masked/tram_1762_v1_f0181_masked.jpg"),  # clean (expected NO)
    ("data/reference/tram_1762_v1_f0227_masked_reference.jpg",
     "data/anomalies/tram_1762_v3_f0219_masked.jpg"),  # objects + real person
]
TASK3_IMAGES = list(TASK1_IMAGES)                  # bbox OUTPUT (vlm_03 prompt)
TASK3Z_IMAGES = [                                  # official zones-as-INPUT task
    "data/anomalies/tram_1762_v2_f0037_masked.jpg",
    "data/masked/tram_1762_v1_f0181_masked.jpg",   # clean (expected NO)
]

ZONES_JSON = "benchmark/zones_tram_1762.json"
CSV_PATH = "results/grid_results.csv"

# Official zones-of-interest prompt, translated template from the spreadsheet's
# 'Example Prompts' sheet (Task 3 - "Zones of interest - bounding boxes").
ZONES_PROMPT_TEMPLATE = """You are a tram interior diagnostic system. Analyze the image and assess the condition of the vehicle. During the analysis focus PRIMARILY on the zones of interest (bounding boxes) defined below. The box coordinates have been automatically converted to the normalized format [ymin, xmin, ymax, xmax] on a 0-1000 scale.
ZONES OF INTEREST TO ANALYZE:
{zones_text}
Answer EXACTLY in this structure (in English):
GRAFFITI: YES/NO
VANDALISM: YES/NO
FORGOTTEN OBJECT(S): YES/NO
DESCRIPTION: where and what (specify exactly in which zone/seat the anomaly is)
SEVERITY: 1-5 (1 = no problem, 5 = major damage)
PEOPLE: YES/NO + COUNT
Be brief and objective, but very precise. No introduction, no conclusion."""

# =============================================================================

COLUMNS = ["ID", "Student", "Date", "Task", "Model", "Test image",
           "Reference image", "Bbox source", "Prompt", "Model response",
           "Anomaly (YES/NO)", "Anomaly type", "Correctness", "Rating (1-5)",
           "Inference time [s]", "Note"]


def _p(path):
    p = Path(path)
    return str(p if p.is_absolute() else REPO_ROOT / p)


def zones_prompt():
    data = json.loads(Path(_p(ZONES_JSON)).read_text(encoding="utf-8"))
    w, h = data["image_width"], data["image_height"]
    lines = []
    for z in data["zones"]:
        ymin = round(z["y"] / h * 1000)
        xmin = round(z["x"] / w * 1000)
        ymax = round((z["y"] + z["h"]) / h * 1000)
        xmax = round((z["x"] + z["w"]) / w * 1000)
        lines.append(f"- {z['label']}: [{ymin}, {xmin}, {ymax}, {xmax}]")
    return ZONES_PROMPT_TEMPLATE.format(zones_text="\n".join(lines))


def chat(model, prompt, images):
    t0 = time.time()
    response = ollama.chat(
        model=model,
        messages=[{"role": "user", "content": prompt,
                   "images": [_p(i) for i in images]}],
        think=False,
        options={"num_ctx": v1.NUM_CTX, "num_predict": v1.NUM_PREDICT,
                 "temperature": v1.TEMPERATURE},
    )
    seconds = time.time() - t0
    message = response.get("message", {})
    text = (message.get("content", "") or "").strip() or \
           (message.get("thinking", "") or "").strip()
    return text, seconds


def yes_no_and_type(task, text):
    """Heuristic Anomaly YES/NO + type from a structured response (or from the
    detection count for the JSON task). The human still fills Correctness."""
    if task == "3":
        dets = v3.parse_json(text)
        types = sorted({str(d.get("label", "?")) for d in dets})
        return ("YES" if dets else "NO"), ", ".join(types)
    found = []
    for field in ("GRAFFITI", "VANDALISM", "FORGOTTEN OBJECT"):
        if re.search(rf"{field}[^:\n]*:\s*yes", text, re.IGNORECASE):
            found.append(field.lower())
    return ("YES" if found else "NO"), ", ".join(found)


def build_calls():
    calls = []   # (task, prompt, images:[...], test_image, reference, bbox_src)
    for img in TASK1_IMAGES:
        calls.append(("1", v1.PROMPT, [img], img, "", ""))
    for ref, insp in TASK2_PAIRS:
        calls.append(("2", v2.PROMPT, [ref, insp], insp, ref, ""))
    for img in TASK3_IMAGES:
        calls.append(("3", v3.PROMPT, [img], img, "", ""))
    zp = zones_prompt()
    for img in TASK3Z_IMAGES:
        calls.append(("3z", zp, [img], img, "", Path(ZONES_JSON).name))
    return calls


def row_key(model, task, test_image, prompt):
    h = hashlib.sha1(prompt.encode("utf-8")).hexdigest()[:10]
    return f"{model}|{task}|{Path(test_image).name}|{h}"


def load_done(csv_path):
    done = set()
    max_id = -1
    p = Path(csv_path)
    if not p.exists():
        return done, max_id
    with open(p, newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            done.add(row_key(row["Model"], str(row["Task"]),
                             row["Test image"], row["Prompt"]))
            try:
                max_id = max(max_id, int(float(row["ID"])))
            except (ValueError, KeyError):
                pass
    return done, max_id


def append_csv(csv_path, row):
    p = Path(_p(csv_path))
    p.parent.mkdir(parents=True, exist_ok=True)
    new = not p.exists()
    with open(p, "a", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=COLUMNS)
        if new:
            w.writeheader()
        w.writerow(row)


def append_xlsx(xlsx_path, rows):
    import openpyxl
    src = Path(xlsx_path)
    backup = src.with_suffix(f".bak-{time.strftime('%Y%m%d-%H%M%S')}.xlsx")
    shutil.copy2(src, backup)
    wb = openpyxl.load_workbook(src)
    ws = wb["Results"]
    first_free = ws.max_row + 1
    for r in range(2, ws.max_row + 1):        # find first row with an empty ID
        if ws.cell(row=r, column=1).value is None:
            first_free = r
            break
    for i, row in enumerate(rows):
        for c, col in enumerate(COLUMNS, start=1):
            ws.cell(row=first_free + i, column=c, value=row[col])
    wb.save(src)
    print(f"xlsx: appended {len(rows)} row(s) at row {first_free} "
          f"(backup: {backup.name})")


def main():
    ap = argparse.ArgumentParser(description="ARSI VLM benchmark grid driver")
    ap.add_argument("--models", default=",".join(MODELS),
                    help="comma-separated Ollama model names")
    ap.add_argument("--tasks", default="1,2,3,3z",
                    help="subset of 1,2,3,3z")
    ap.add_argument("--images", default="",
                    help="only test images whose path contains this substring")
    ap.add_argument("--csv", default=CSV_PATH)
    ap.add_argument("--to-xlsx", default=None, metavar="XLSX",
                    help="also append this run's rows into the workbook "
                         "(a .bak copy is made first)")
    ap.add_argument("--dry-run", action="store_true",
                    help="print the call plan and exit")
    args = ap.parse_args()

    models = [m.strip() for m in args.models.split(",") if m.strip()]
    tasks = {t.strip() for t in args.tasks.split(",")}
    calls = [c for c in build_calls()
             if c[0] in tasks and args.images in c[3]]

    done, max_id = load_done(_p(args.csv))
    plan = [(model, c) for model in models for c in calls
            if row_key(model, c[0], c[3], c[1]) not in done]
    print(f"{len(plan)} call(s) to run ({len(done)} already in "
          f"{args.csv}); models: {', '.join(models)}")
    if args.dry_run:
        for model, (task, _p_, imgs, test, ref, bbox) in plan:
            print(f"  task {task:<3} {model:<45} {Path(test).name}")
        return

    installed = None
    try:
        data = ollama.list()
        installed = {getattr(m, "model", None) or m.get("model")
                     for m in (getattr(data, "models", None) or data["models"])}
    except Exception:
        pass

    new_rows = []
    next_id = max_id + 1
    for n, (model, (task, prompt, images, test, ref, bbox)) in enumerate(plan, 1):
        if installed is not None and model not in installed \
                and f"{model}:latest" not in installed:
            print(f"[{n}/{len(plan)}] SKIP {model} (not installed - "
                  f"ollama pull {model})")
            continue
        print(f"[{n}/{len(plan)}] task {task:<3} {model:<45} "
              f"{Path(test).name} ...", end="", flush=True)
        try:
            text, seconds = chat(model, prompt, images)
        except Exception as exc:
            print(f" ERROR ({type(exc).__name__}: {exc})")
            continue
        yes, typ = yes_no_and_type(task, text)
        print(f" {seconds:6.1f}s  anomaly={yes}")
        row = {"ID": next_id, "Student": STUDENT,
               "Date": date.today().isoformat(), "Task": task, "Model": model,
               "Test image": Path(test).name,
               "Reference image": Path(ref).name if ref else "",
               "Bbox source": bbox, "Prompt": prompt, "Model response": text,
               "Anomaly (YES/NO)": yes, "Anomaly type": typ,
               "Correctness": "", "Rating (1-5)": "",
               "Inference time [s]": round(seconds, 1), "Note": ""}
        append_csv(args.csv, row)
        new_rows.append(row)
        next_id += 1

    if args.to_xlsx and new_rows:
        append_xlsx(args.to_xlsx, new_rows)
    print(f"\nDone: {len(new_rows)} new row(s) -> {args.csv}")


if __name__ == "__main__":
    main()

"""Review layer: validation rules, metric computation (supervisor's
any-miss-is-FN frame rule), and the API round-trip including the xlsx export."""
import pytest

from app.backend.review import (ReviewError, compute_metrics, validate_review)

from test_backend import (VALID_VLM01, api, app_image,  # noqa: F401
                          wait_job)


def results_fixture():
    return {"frames": [
        {"frame_id": "fA", "status": "ok", "anomaly": True,
         "detections": [{"label": "phone", "type": "object", "bbox": [1, 2, 30, 40]},
                        {"label": "tag", "type": "graffiti", "bbox": [5, 5, 20, 20]}]},
        {"frame_id": "fB", "status": "ok", "anomaly": False, "detections": []},
        {"frame_id": "fC", "status": "ok", "anomaly": False, "detections": []},
        {"frame_id": "fD", "status": "failed", "anomaly": None, "detections": []},
        {"frame_id": "fE", "status": "ok", "anomaly": True,
         "detections": [{"label": "bag", "type": "object", "bbox": [0, 0, 9, 9]}]},
    ]}


# ---------------------------------------------------------------- validation

def test_validate_rejects_garbage():
    res = results_fixture()
    with pytest.raises(ReviewError, match="unknown frame_id"):
        validate_review(res, {"nope": {}})
    with pytest.raises(ReviewError, match="out of range"):
        validate_review(res, {"fA": {"verdicts": {"7": "tp"}}})
    with pytest.raises(ReviewError, match="verdict must be"):
        validate_review(res, {"fA": {"verdicts": {"0": "maybe"}}})
    with pytest.raises(ReviewError, match="not an index"):
        validate_review(res, {"fA": {"verdicts": {"x": "tp"}}})
    with pytest.raises(ReviewError, match="needs a label"):
        validate_review(res, {"fB": {"missed": [{"bbox": [0, 0, 5, 5]}]}})
    with pytest.raises(ReviewError, match="empty or inverted"):
        validate_review(res, {"fB": {"missed": [{"bbox": [10, 10, 5, 20],
                                                 "label": "x"}]}})
    with pytest.raises(ReviewError, match="cannot confirm"):
        validate_review(res, {"fA": {"verdicts": {"0": "tp"}, "done": True}})


def test_validate_normalizes():
    res = results_fixture()
    out = validate_review(res, {
        "fA": {"verdicts": {"0": "tp", "1": "fp"}, "done": True},
        "fB": {"missed": [{"bbox": [1.4, 2.6, 30, 44], "label": "  phone ",
                           "type": "weird"}], "done": True},
        "fC": {"verdicts": {}, "missed": [], "done": False},   # empty -> dropped
    })
    assert set(out) == {"fA", "fB"}
    m = out["fB"]["missed"][0]
    assert m == {"bbox": [1, 3, 30, 44], "label": "phone", "type": "unknown"}


# ---------------------------------------------------------------- metrics

def test_metrics_supervisor_rules():
    res = results_fixture()
    review = {"frames": {
        # anomalous frame, one TP one FP -> frame TP (fp noted at object level)
        "fA": {"verdicts": {"0": "tp", "1": "fp"}, "missed": [], "done": True},
        # clean-predicted frame with a missed object -> frame FN
        "fB": {"verdicts": {}, "done": True,
               "missed": [{"bbox": [0, 0, 9, 9], "label": "wallet",
                           "type": "object"}]},
        # clean-predicted, confirmed clean -> TN
        "fC": {"verdicts": {}, "missed": [], "done": True},
        # failed frame acknowledged -> excluded from the confusion matrix
        "fD": {"verdicts": {}, "missed": [], "done": True},
        # verdicts set but NOT confirmed -> excluded everywhere
        "fE": {"verdicts": {"0": "tp"}, "missed": [], "done": False},
    }}
    m = compute_metrics(res, review)
    assert m["progress"] == {"n_frames": 5, "n_done": 4, "n_failed": 1}
    assert m["objects"] == {"tp": 1, "fp": 1, "fn": 1, "precision": 0.5,
                            "recall": 0.5, "f1": 0.5}
    fr = m["frames"]
    assert (fr["TP"], fr["FP"], fr["TN"], fr["FN"]) == (1, 0, 1, 1)
    assert fr["n_scored"] == 3 and fr["accuracy"] == 0.667
    assert m["correctness"] == {"fA": "TP", "fB": "FN", "fC": "TN"}
    assert m["per_type"]["object"] == {"tp": 1, "fp": 0, "fn": 1, "recall": 0.5}
    assert m["per_type"]["graffiti"]["fp"] == 1


def test_metrics_all_fp_frame_scores_fp():
    res = {"frames": [{"frame_id": "f1", "status": "ok", "anomaly": True,
                       "detections": [{"label": "ghost", "type": "object",
                                       "bbox": [0, 0, 5, 5]}]}]}
    review = {"frames": {"f1": {"verdicts": {"0": "fp"}, "missed": [],
                                "done": True}}}
    m = compute_metrics(res, review)
    assert m["correctness"] == {"f1": "FP"} and m["frames"]["specificity"] == 0.0


# ---------------------------------------------------------------- API round-trip

def test_review_api_roundtrip(api):
    client, _ = api(replies=[VALID_VLM01])
    r = client.post("/api/jobs", json={"script": "vlm_01",
                                       "frames": [str(app_image("rev.jpg"))],
                                       "model": "qwen3-vl:8b-instruct"})
    job_id = r.json()["job_id"]
    wait_job(client, job_id)

    r = client.get(f"/api/jobs/{job_id}/review")
    assert r.status_code == 200
    body = r.json()
    assert body["review"]["frames"] == {}
    assert body["metrics"]["progress"]["n_done"] == 0

    frame_id = client.get(f"/api/jobs/{job_id}").json()["frames"][0]["frame_id"]
    bad = {"frames": {frame_id: {"verdicts": {"9": "tp"}}}}
    assert client.put(f"/api/jobs/{job_id}/review", json=bad).status_code == 400

    good = {"frames": {frame_id: {
        "verdicts": {"0": "tp"}, "done": True,
        "missed": [{"bbox": [10, 10, 60, 60], "label": "wallet",
                    "type": "object"}]}}}
    r = client.put(f"/api/jobs/{job_id}/review", json=good)
    assert r.status_code == 200
    m = r.json()["metrics"]
    assert m["objects"]["tp"] == 1 and m["objects"]["fn"] == 1
    assert m["correctness"][frame_id] == "FN"          # any miss -> FN

    # persisted: a fresh GET returns the same review
    again = client.get(f"/api/jobs/{job_id}/review").json()
    assert again["review"]["frames"][frame_id]["verdicts"] == {"0": "tp"}

    # xlsx export now carries the review columns and sheet
    from io import BytesIO
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(client.get(f"/api/jobs/{job_id}/export.xlsx").content))
    assert "Review" in wb.sheetnames
    header = [c.value for c in wb["Frames"][1]]
    assert "Correctness" in header
    row = [c.value for c in wb["Frames"][2]]
    assert row[header.index("Correctness")] == "FN"
    assert row[header.index("Reviewed")] == "yes"

    assert client.get("/api/jobs/nope/review").status_code == 404


def test_review_pure_helpers_dont_need_api():
    assert compute_metrics({"frames": []}, {"frames": {}})["frames"]["accuracy"] is None


# ---------------------------------------------------------------- labels + lora screens

def make_reviewed_vlm05_job(job_id="fakejob-lora01"):
    """A saved vlm_05-style job + confirmed review, crafted directly in
    JOBS_DIR — what the Labels/LoRA endpoints and the exporter consume."""
    import json
    from arsi_core.runner import JOBS_DIR
    ref = app_image("lora_ref.jpg", size=(400, 300))
    insp = app_image("lora_insp.jpg", size=(400, 300), color=(110, 110, 110))
    jdir = JOBS_DIR / job_id
    jdir.mkdir(parents=True, exist_ok=True)
    results = {
        "job_id": job_id, "status": "completed",
        "config": {"script": "vlm_05", "model": "m", "prompt_name": "conservative",
                   "prompt": "p", "reference": str(ref), "mask": None,
                   "n_frames": 1, "params": {}},
        "started": "2026-07-19T00:00:00+00:00", "finished": "",
        "frames": [{"frame_id": "f0001", "image": str(insp), "status": "ok",
                    "attempts": 1, "seconds": 1.0, "anomaly": True,
                    "detections": [
                        {"label": "phone", "type": "object", "bbox": [40, 40, 90, 80],
                         "zone": None, "severity": None, "confidence": None,
                         "channel": "photo"},
                        {"label": "ghost", "type": "object", "bbox": [200, 100, 260, 150],
                         "zone": None, "severity": None, "confidence": None,
                         "channel": "photo"}],
                    "raw_response": "", "error": None}],
        "summary": {"n_frames": 1, "n_ok": 1, "n_anomalous": 1, "n_failed": 0,
                    "wall_seconds": 1.0}}
    with open(jdir / "results.json", "w") as fh:
        json.dump(results, fh)
    review = {"job_id": job_id, "updated": "2026-07-19T00:00:00+00:00",
              "frames": {"f0001": {
                  "verdicts": {"0": "tp", "1": "fp"},
                  "missed": [{"bbox": [300, 200, 360, 260], "label": "wallet",
                              "type": "object"}],
                  "done": True}}}
    with open(jdir / "review.json", "w") as fh:
        json.dump(review, fh)
    return jdir


def test_reviews_index_lora_status_and_export(api):
    client, _ = api()
    jdir = make_reviewed_vlm05_job()
    try:
        rows = client.get("/api/reviews").json()["reviews"]
        row = next(r for r in rows if r["job_id"] == jdir.name)
        assert row["metrics"]["objects"]["tp"] == 1
        assert row["export"] == {"yes": 2, "no": 1, "skipped_no_bbox": 0,
                                 "exportable": True, "samples": 3}

        st = client.get("/api/lora/status").json()
        assert st["aggregate"]["samples"] >= 3
        assert any(j["job_id"] == jdir.name for j in st["per_job"])

        r = client.post("/api/lora/export", json={})
        assert r.status_code == 200, r.text
        stats = r.json()["stats"]
        assert stats["total"] >= 3 and stats["yes"] >= 2 and stats["no"] >= 1
        assert client.get("/api/lora/status").json()["last_export"]["total"] >= 3

        assert client.delete(f"/api/jobs/{jdir.name}/review").json() == {"ok": True}
        assert client.delete(f"/api/jobs/{jdir.name}/review").status_code == 404
        assert all(r["job_id"] != jdir.name
                   for r in client.get("/api/reviews").json()["reviews"])
    finally:
        import shutil
        shutil.rmtree(jdir, ignore_errors=True)
